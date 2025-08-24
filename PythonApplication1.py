# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os

R = 8.314  # J/(mol*K) gaz sabiti

def read_element_properties(file_path):
    df = pd.read_excel(file_path)
    atomic_number_map = list(df['Atomic Number'])
    return df, atomic_number_map

def read_miedema_matrix(file_path):
    matrix = pd.read_excel(file_path, index_col=0)
    return matrix

def parse_key(key):
    key = key.strip()
    try:
        return int(key)
    except ValueError:
        return int(key, 16)

def get_elements(atomic_number_map):
    print("# Elementlerin atom numarası ve sembolü ile giriş yapılacaktır.")
    print("# Aşağıda kullanılabilir anahtarlar gösterilmiştir:")
    for key in atomic_number_map:
        print(f"# {key}")
    print("Yüksek entropili alaşım elementlerinin anahtarlarını girin (virgülle ayrılmış, örn: 24Cr,27Co,40Zr):")
    elements_input = input("Anahtarlar: ")
    selected_keys = [key.strip() for key in elements_input.split(",")]
    for key in selected_keys:
        if key not in atomic_number_map:
            print(f"Uyarı: {key} anahtarı listede yok!")
    return selected_keys

def get_compositions(selected_keys):
    print("\nSeçtiğiniz elementler için yüzde kompozisyonları giriniz (toplam 100 olmalı):")
    compositions = {}
    while True:
        for key in selected_keys:
            while True:
                try:
                    percent = float(input(f"{key} için yüzde: "))
                    if percent < 0 or percent > 100:
                        print("Yüzde 0 ile 100 arasında olmalı.")
                        continue
                    compositions[key] = percent
                    break
                except ValueError:
                    print("Lütfen geçerli bir sayı girin.")
        total = sum(compositions.values())
        if abs(total - 100) < 1e-6:
            break
        else:
            print(f"Toplam yüzde {total}. Lütfen tekrar giriniz (toplam 100 olmalı).")
            compositions.clear()
    print("\nGirilen kompozisyonlar:")
    for key, val in compositions.items():
        print(f"{key}: %{val}")
    return compositions

def filter_elements(df, selected_keys):
    return df[df['Atomic Number'].isin(selected_keys)]

def print_table(filtered_df):
    columns = [
        "Atomic Number", "Element", "Atomic Weight", "Density", "Atomic Radius",
        "Melting Point", "Lattice Constant", "Vickers Hardness",
        "Young's Modulus", "Thermal Neutron Absorption Cross-section",
        "Valence Electron Concentration", "Pauling Electronegativity"
    ]
    table = filtered_df[columns]
    print("\nSeçilen elementlerin özellik tablosu:\n")
    print(table.to_string(index=False))

def calculate_mole_fractions(compositions):
    total = sum(compositions.values())
    return {key: val / total for key, val in compositions.items()}

def calculate_entropy(mole_fractions):
    S_mix = -R * sum(ci * np.log(ci) for ci in mole_fractions.values())
    return S_mix

def calculate_binary_enthalpy_dict(matrix, selected_keys):
    binary_enthalpies_kjmol = {}
    for i in selected_keys:
        for j in selected_keys:
            if i != j:
                try:
                    value = matrix.loc[i, j]
                    binary_enthalpies_kjmol[(i, j)] = value
                except KeyError:
                    continue
    return binary_enthalpies_kjmol

def calculate_mixture_enthalpy(selected_keys, mole_fractions, binary_enthalpies_kjmol):
    delta_H_kjmol = 0.0
    print("\nKatkılar (kJ/mol):")
    for i in range(len(selected_keys)):
        for j in range(i + 1, len(selected_keys)):
            e1 = selected_keys[i]
            e2 = selected_keys[j]
            pair = (e1, e2)
            reverse_pair = (e2, e1)
            H_ij = binary_enthalpies_kjmol.get(pair) or binary_enthalpies_kjmol.get(reverse_pair)
            if H_ij is not None and not np.isnan(H_ij):
                contrib = 2 * mole_fractions[e1] * mole_fractions[e2] * H_ij
                delta_H_kjmol += contrib
                print(f"{e1}-{e2}: 2 * {mole_fractions[e1]:.4f} * {mole_fractions[e2]:.4f} * ({H_ij:.4f}) = {contrib:.4f} kJ/mol")
    return delta_H_kjmol

def calculate_gibbs(H_mix, S_mix, T):
    return H_mix * 1000 - T * S_mix  # H_mix kJ/mol ise J/mol'a çevir

def calculate_melting_temperature(mole_fractions, filtered_df):
    Tm = 0.0
    for key, ci in mole_fractions.items():
        Tm_i = float(filtered_df[filtered_df['Atomic Number'] == key]['Melting Point'].values[0])
        Tm += ci * Tm_i
    return Tm

def calculate_max_entropy_ratio(Tm, S_mix, delta_H_kjmol):
    delta_H_jmol = abs(delta_H_kjmol * 1000)
    omega = (Tm * S_mix) / delta_H_jmol if delta_H_jmol != 0 else float('inf')
    return omega

def calculate_atomic_size_difference(mole_fractions, filtered_df):
    r_bar = sum(
        mole_fractions[key] * float(filtered_df[filtered_df['Atomic Number'] == key]['Atomic Radius'].values[0])
        for key in mole_fractions
    )
    delta = 0.0
    for key, ci in mole_fractions.items():
        ri = float(filtered_df[filtered_df['Atomic Number'] == key]['Atomic Radius'].values[0])
        delta += ci * ((1 - ri / r_bar) ** 2)
    delta = (delta ** 0.5) * 100  # % cinsinden
    return delta, r_bar

def calculate_vec(mole_fractions, filtered_df):
    vec = 0.0
    for key, ci in mole_fractions.items():
        vec_i = float(filtered_df[filtered_df['Atomic Number'] == key]['Valence Electron Concentration'].values[0])
        vec += ci * vec_i
    return vec

def calculate_electronegativity_difference(mole_fractions, filtered_df):
    chi_values = {}
    for key in mole_fractions:
        chi = float(filtered_df[filtered_df['Atomic Number'] == key]['Pauling Electronegativity'].values[0])
        chi_values[key] = chi
    chi_bar = sum(mole_fractions[key] * chi_values[key] for key in mole_fractions)
    delta_chi = (sum(mole_fractions[key] * (chi_values[key] - chi_bar) ** 2 for key in mole_fractions)) ** 0.5
    delta_chi_percent = (delta_chi / chi_bar) * 100 if chi_bar != 0 else 0
    return chi_bar, delta_chi, delta_chi_percent

def export_results_to_excel(filename, results_dict):
    df = pd.DataFrame(list(results_dict.items()), columns=["Parametre", "Değer"])
    df.to_excel(filename, index=False, sheet_name="Alaşım Özeti")
    wb = load_workbook(filename)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        row[0].alignment = Alignment(horizontal="left")
        row[1].alignment = Alignment(horizontal="right")
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 22
    wb.save(filename)
    print(f"\nSonuçlar '{filename}' dosyasına kaydedildi.")
    try:
        os.startfile(filename)
        print(f"Excel dosyası otomatik olarak açıldı: {filename}")
    except Exception as e:
        print(f"Excel dosyası açılamadı: {e}")

def create_lammps_input(selected_keys, compositions, filtered_df, output_dir="."):
    element_names = [str(filtered_df[filtered_df['Atomic Number'] == k]['Element'].values[0]) for k in selected_keys]
    file_name = f"in_{'_'.join(element_names)}.lmp"
    file_path = os.path.join(output_dir, file_name)

    a0 = np.mean([
        float(filtered_df[filtered_df['Atomic Number'] == k]['Lattice Constant'].values[0])
        for k in selected_keys
    ])
    masses = [
        float(filtered_df[filtered_df['Atomic Number'] == k]['Atomic Weight'].values[0])
        for k in selected_keys
    ]
    fractions = [compositions[k]/100 for k in selected_keys]

    with open(file_path, "w") as f:
        f.write("clear\n")
        f.write("#------------------INITIALIZATION--------------------------\n")
        f.write("units metal\n")
        f.write("dimension 3\n")
        f.write("boundary p p p\n")
        f.write("atom_style atomic\n\n")
        f.write("#------------------ALLOYING-------------------------------\n")
        f.write(f"#Alloy: {'-'.join(element_names)} HEA\n\n")
        f.write("#------------------SIMULATION CELL------------------------\n")
        f.write("region whole block -50 50 -150 150 -50 50 units box\n")
        f.write(f"create_box {len(selected_keys)} whole\n\n")
        f.write("#------------------ATOM DEFINITION------------------------\n")
        f.write("region nw cylinder y 0 0 50 INF units box\n")
        f.write(f"lattice bcc {a0:.2f} orient x 1 0 0 orient y 0 1 0 orient z 0 0 1\n")
        f.write("create_atoms 1 region nw\n")
        for i, (frac, name) in enumerate(zip(fractions, element_names)):
            f.write(f"set type {i+1} type/fraction {i+1} {frac:.2f} 1234567 # {name}\n")
        f.write("\n#------------------PAIR COEFF-----------------------------\n")
        f.write("pair_style lj/cut 10\n")
        f.write("pair_coeff * * 0 0\n\n")
        f.write("#------------------MASSES---------------------------------\n")
        for i, (mass, name) in enumerate(zip(masses, element_names)):
            f.write(f"mass {i+1} {mass} # {name}\n")
        f.write("\n#Write alloying Elements as comment\n")
        f.write("delete_atoms overlap 0.3 all all\n\n")
        f.write("#------------------EXPORT DATA FILE-----------------------\n")
        f.write("write_data structure.lmpdat noceoff\n")
    print(f"LAMMPS input dosyası '{file_name}' oluşturuldu.")

if __name__ == "__main__":
    properties_file = "element_properties.xlsx"
    miedema_file = "miedema_matrix.xlsx"

    df, atomic_number_map = read_element_properties(properties_file)
    miedema_matrix = read_miedema_matrix(miedema_file)
    selected_keys = get_elements(atomic_number_map)
    compositions = get_compositions(selected_keys)
    filtered_df = filter_elements(df, selected_keys)
    print_table(filtered_df)

    mole_fractions = calculate_mole_fractions(compositions)
    print("\nMol oranları:")
    for key, val in mole_fractions.items():
        print(f"{key}: {val:.4f}")

    S_mix = calculate_entropy(mole_fractions)
    print(f"\nKarışım Entropisi (ΔS_mix): {S_mix:.4f} J/mol·K")

    binary_enthalpies_kjmol = calculate_binary_enthalpy_dict(miedema_matrix, selected_keys)
    delta_H_kjmol = calculate_mixture_enthalpy(selected_keys, mole_fractions, binary_enthalpies_kjmol)
    print(f"\nKarışım Entalpisi (ΔH_mix): {delta_H_kjmol:.4f} kJ/mol")

    T = float(input("\nHesaplama için sıcaklık değerini giriniz (K): "))
    G_mix = calculate_gibbs(delta_H_kjmol, S_mix, T)
    print(f"\nGibbs Serbest Enerjisi (ΔG_mix): {G_mix:.4f} J/mol")

    if G_mix < 0:
        print("ΔG_mix < 0: Alaşımın tek fazlı katı çözelti oluşturma olasılığı yüksek.")
    else:
        print("ΔG_mix > 0: Faz ayrımı veya ara fazlar oluşabilir.")
    print("Ancak, aşağıdaki kriterlerin kontrolü yapılması gereklidir.")
    print("\n--- Alaşımın Termodinamik Kriterleri ---")
    enthalpy_ok = (-10.0001 <= delta_H_kjmol <= 5.0001)
    entropy_ok = (S_mix >= 1.5 * R)

    print(f"ΔH_mix: {delta_H_kjmol:.2f} kJ/mol {'(Uygun)' if enthalpy_ok else '(Uygun Değil)'}")
    print(f"ΔS_mix: {S_mix:.2f} J/mol·K {'(Uygun)' if entropy_ok else '(Uygun Değil)'}")
    print("Kriterler: ΔH_mix −10 ile +5 kJ/mol arasında ve ΔS_mix ≥ 1.5 R olmalıdır.")
    print("Burada konfigurasyonel entropi (ΔS_mix) büyük olmalı, entalpi (ΔH_mix) ise çok negatif olmamalıdır.")
    print("Genel olarak ΔH_mix −10 ile +5 kJ/mol arasında ve karışım entropisi≥1.5 R olmalıdır.")

    if enthalpy_ok and entropy_ok:
        print("Bu alaşım, tek fazlı katı çözelti oluşturma açısından termodinamik olarak uygundur.")
    else:
        print("Bu alaşım, tek fazlı katı çözelti oluşturma açısından termodinamik olarak uygun değildir.")

    vec = calculate_vec(mole_fractions, filtered_df)
    print(f"\nAlaşımın Valence Electron Concentration (VEC) değeri: {vec:.2f}")

    if vec < 6.87:
        faz = "BCC (VEC < 6.87)"
        print("Beklenen faz: BCC (VEC < 6.87)")
    elif 6.87 <= vec < 8.00:
        faz = "BCC + FCC (6.87 ≤ VEC < 8.00)"
        print("Beklenen faz: BCC + FCC (6.87 ≤ VEC < 8.00)")
    else:
        faz = "FCC (VEC ≥ 8.00)"
        print("Beklenen faz: FCC (VEC ≥ 8.00)")

    Tm = calculate_melting_temperature(mole_fractions, filtered_df)
    print(f"\nAlaşımın ortalama ergime sıcaklığı (Tₘ): {Tm:.2f} K")

    omega = calculate_max_entropy_ratio(Tm, S_mix, delta_H_kjmol)
    print(f"\nMaksimum entropi oranı (Ω): {omega:.2f}")
    if omega >= 1.1:
        omega_kriter = "Uygun"
        print("Ω ≥ 1.1: Entropi entalpiden üstündür, tek fazlı çözelti oluşumu için yeterlidir.")
    else:
        omega_kriter = "Uygun Değil"
        print("Ω < 1.1: Entropi entalpiden üstün değildir, tek fazlı çözelti oluşumu için yeterli değildir.")

    delta, r_bar = calculate_atomic_size_difference(mole_fractions, filtered_df)
    print(f"\nAlaşımın ortalama atomik yarıçapı: {r_bar:.4f} Å")
    print(f"Atomik boyut farkı (δ): {delta:.2f} %")

    print("Kararlılık kriteri: δ ≤ 6.6% ise tek fazlı katı çözelti oluşumu olasılığı yüksektir.")
    if delta <= 6.6:
        delta_kriter = "Uygun"
        print("δ ≤ 6.6%: Tek fazlı katı çözelti oluşumu olasılığı yüksektir.")
    else:
        delta_kriter = "Uygun Değil"
        print("δ > 6.6%: Kristal yapı bozulabilir, faz ayrımı veya intermetalik fazlar oluşabilir.")

    chi_bar, delta_chi, delta_chi_percent = calculate_electronegativity_difference(mole_fractions, filtered_df)
    print(f"\nAlaşımın ortalama Pauling elektronegatifliği (χ̄): {chi_bar:.3f}")
    print(f"Elektronegatiflik farkı (Δχ): {delta_chi:.3f}")
    print(f"Yüzdesel elektronegatiflik farkı (Δχ%): {delta_chi_percent:.2f} %")

    if delta_chi_percent <= 8:
        chi_yorum = "Küçük/ılımlı (≤8%): Tek faz katı çözelti olasılığını artırabilir, ama garanti değildir."
        print("Δχ% küçük/ılımlı (≤8%): Tek faz katı çözelti olasılığını artırabilir, ama garanti değildir.")
    elif delta_chi_percent >= 10:
        chi_yorum = "Yüksek (≥10%): İntermetalik/ayrışma riski artar; yine de refrakter BCC ailesinde istisnalar görülebilir."
        print("Δχ% yüksek (≥10%): İntermetalik/ayrışma riski artar; yine de refrakter BCC ailesinde istisnalar görülebilir.")
    else:
        chi_yorum = "Orta (8-10%): Faz davranışı için diğer kriterlerle birlikte değerlendirilmelidir."
        print("Δχ% orta aralıkta (8-10%): Faz davranışı için diğer kriterlerle birlikte değerlendirilmelidir.")

    results = {
        "Valence Electron Concentration (VEC)": f"{vec:.2f}",
        "Beklenen Faz": faz,
        "Karışım Entalpisi (ΔH_mix) [kJ/mol]": f"{delta_H_kjmol:.4f}",
        "Karışım Entropisi (ΔS_mix) [J/mol·K]": f"{S_mix:.4f}",
        "Gibbs Serbest Enerjisi (ΔG_mix) [J/mol]": f"{G_mix:.4f}",
        "Ortalama Ergime Sıcaklığı (Tm) [K]": f"{Tm:.2f}",
        "Maksimum Entropi Oranı (Ω)": f"{omega:.2f}",
        "Ω Kriteri": omega_kriter,
        "Ortalama Atomik Yarıçap (r̄) [Å]": f"{r_bar:.4f}",
        "Atomik Boyut Farkı (δ) [%]": f"{delta:.2f}",
        "δ Kriteri": delta_kriter,
        "Ortalama Pauling Elektronegatifliği (χ̄)": f"{chi_bar:.3f}",
        "Elektronegatiflik Farkı (Δχ)": f"{delta_chi:.3f}",
        "Yüzdesel Elektronegatiflik Farkı (Δχ%)": f"{delta_chi_percent:.2f} %",
        "Δχ% Yorumu": chi_yorum,
        "ΔH_mix Kriteri": "Uygun" if enthalpy_ok else "Uygun Değil",
        "ΔS_mix Kriteri": "Uygun" if entropy_ok else "Uygun Değil",
        "ΔG_mix Yorumu": (
            "ΔG_mix < 0: Tek fazlı katı çözelti olasılığı yüksek."
            if G_mix < 0 else
            "ΔG_mix > 0: Faz ayrımı veya ara fazlar oluşabilir."
        ),
    }

    export_results_to_excel("alasim_ozet.xlsx", results)

    karar = input("\nAlaşım uygun mu? (E/h): ")
    if karar.strip().lower() in ["e", "evet", "y"]:
        create_lammps_input(selected_keys, compositions, filtered_df)